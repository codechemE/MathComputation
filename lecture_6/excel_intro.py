from openpyxl import load_workbook

wb = load_workbook(filename="Students.xlsx")
print(wb.sheetnames)
sheet1 = wb['Sheet1']
# print(tuple(sheet.rows))
# print(tuple(sheet.columns))
for sheet in wb:
    print(sheet.title)

name = "Jefim Naljotov"
course = "CME502"
department = "CME"
semester = 2

sheet1['A7'] = name
sheet1['B7'] = course
sheet1['C7'] = department
sheet1['D7'] = semester

wb.save("copy_Students.xlsx")